import shutil
from openpyxl import load_workbook, Workbook

#Create duplicate files
# for i in range(2, 6):       #Creates DATA2.xlsx to DATA5.xlsx 
#     shutil.copyfile("DATA1.xlsx", f"DATA{i}.xlsx")

#Create summary workbook
summary = Workbook()
sheet = summary.active
sheet.title = "Summary"

sheet.append(["Row", "Source File", "Total Processing Time"])

#Process all files
for i in range(3):
    filename = f"./dataFolder/data{i + 1}.xlsx"

    workbook =  load_workbook(filename)
    data = workbook.active

    row_num = 1

    for row in data.iter_rows(values_only=True):
        total = float(row[0]) + float(row[1]) + float(row[2])

        sheet.append([row_num, filename, total])
        row_num += 1

#Save summary file
summary.save("SUMMARY.xlsx")

print("Done!")